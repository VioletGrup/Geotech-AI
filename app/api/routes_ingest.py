"""
routes_ingest.py — PDF → template pipeline.

POST /ingest/extract   accepts multiple PDFs, runs extraction + LLM mapping,
                       returns the populated template as JSON (one key per sheet).

GET  /ingest/template  returns the blank template xlsx for download.

POST /ingest/import    accepts the final template xlsx and bulk-writes every
                       sheet to the graph via the existing /nodes bulk endpoints.
"""
from __future__ import annotations
import asyncio, io, json, os, re, threading, time, traceback, uuid
from typing import List

import openpyxl
from fastapi import APIRouter, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse

from app.ingestion.pdf_extractor import extract_pdf, tables_as_text, ExtractedTable
from app.utils.logger import get_logger

logger = get_logger(__name__)
router = APIRouter(prefix="/ingest", tags=["ingest"])

# ── cancellation registry ─────────────────────────────────────────────────────
# Maps job_id → threading.Event.  Set the event to signal the worker to stop.
_cancel_events: dict[str, threading.Event] = {}

# ── LLM call (same env vars as agent.py, direct HTTP — no agent-framework) ───

def _llm_extract(prompt: str, cancel: threading.Event | None = None) -> str:
    """Call whichever LLM is configured.
    Uses llama-3.1-8b-instant for Groq (131k tokens/min).
    Checks cancel event between retries so cancellation is responsive."""
    import requests

    groq_key = os.getenv("GROQ_API_KEY")
    gemini_key = os.getenv("GEMINI_API_KEY")
    openai_key = os.getenv("OPENAI_API_KEY")

    if groq_key:
        url     = "https://api.groq.com/openai/v1/chat/completions"
        headers = {"Authorization": f"Bearer {groq_key}", "Content-Type": "application/json"}
        model   = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")
    elif gemini_key:
        url     = "https://generativelanguage.googleapis.com/v1beta/openai/chat/completions"
        headers = {"Authorization": f"Bearer {gemini_key}", "Content-Type": "application/json"}
        model   = os.getenv("GEMINI_MODEL", "gemini-2.0-flash")
    elif openai_key:
        url     = "https://api.openai.com/v1/chat/completions"
        headers = {"Authorization": f"Bearer {openai_key}", "Content-Type": "application/json"}
        model   = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
    else:
        raise RuntimeError("No LLM API key configured (GROQ_API_KEY / GEMINI_API_KEY / OPENAI_API_KEY)")

    body = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0,
        "max_tokens": 4096,
    }

    wait = 15
    for attempt in range(4):
        if cancel and cancel.is_set():
            raise RuntimeError("Cancelled")
        r = requests.post(url, headers=headers, json=body, timeout=90)
        if r.status_code == 429:
            logger.warning("Rate limited — waiting %ds (attempt %d)", wait, attempt + 1)
            # Sleep in 1s increments so cancel is checked frequently
            for _ in range(wait):
                if cancel and cancel.is_set():
                    raise RuntimeError("Cancelled")
                time.sleep(1)
            wait = min(wait * 2, 120)
            continue
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"]

    raise RuntimeError("Rate limit — retries exhausted")


def _parse_json(raw: str) -> dict | list:
    """Strip markdown fences, extract the first JSON object/array, then parse."""
    # strip code fences
    raw = re.sub(r"```json|```", "", raw).strip()
    # find the first { or [ and the matching last } or ]
    start = next((i for i, c in enumerate(raw) if c in "{["), None)
    if start is None:
        raise ValueError("No JSON object found in LLM response")
    # find the last matching closer
    closer = "}" if raw[start] == "{" else "]"
    end = raw.rfind(closer)
    if end == -1:
        raise ValueError("Unterminated JSON in LLM response")
    raw = raw[start:end+1]
    # replace single-quoted keys/values (common Llama quirk)
    # only do this when standard parse fails
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        # attempt light repair: trailing commas before } or ]
        repaired = re.sub(r",\s*([}\]])", r"\1", raw)
        try:
            return json.loads(repaired)
        except json.JSONDecodeError:
            # last resort: replace Python-style None/True/False
            repaired = repaired.replace("None", "null").replace("True", "true").replace("False", "false")
            return json.loads(repaired)


# ── Schema definition (mirrors Example_Input.xlsx sheets) ────────────────────

SCHEMA = {
    "site": {
        "columns": ["id","name","address","coordinate_system"],
        "required": ["id"],
        "desc": "One row per project site.",
    },
    "zone": {
        "columns": ["id","site_id","site_id","name","pile_drilling",
                    "trackers_4_string","trackers_3_string","trackers_2_string"],
        "required": ["id","site_id"],
        "desc": "Zones / blocks / PCUs within the site. pile_drilling = Pre-Drill | Driven | None (leave None for new sites).",
    },
    "dpsh": {
        "columns": ["id","site_id","zone_id","easting","northing","refusal_depth"],
        "required": ["id"],
        "desc": "DPSH probe refusal depths (m). zone_id filled later from map.",
    },
    "borehole": {
        "columns": ["id","site_id","zone_id","ground_model_id","series","elevation",
                    "total_depth","groundwater_depth"],
        "required": ["id"],
        "desc": "Boreholes. series = BH or SS-BH.",
    },
    "testpit": {
        "columns": ["id","site_id","zone_id","ground_model_id","elevation","total_depth"],
        "required": ["id"],
        "desc": "Trial / test pits.",
    },
    "soil-type": {
        "columns": ["unit_no","origin","unit_name","description"],
        "required": ["unit_no"],
        "desc": "Soil / material vocabulary. unit_no must be unique globally (e.g. 1, 2A, 4D).",
    },
    "ground-model": {
        "columns": ["id"],
        "required": ["id"],
        "desc": "One ground model per borehole or test pit.",
    },
    "ground-layer": {
        "columns": ["id","site_id","ground_model_id","soil_unit_no","start_depth","end_depth"],
        "required": ["id","ground_model_id","start_depth","end_depth"],
        "desc": "Depth bands within a ground model. One row per (layer, soil type) pair.",
    },
    "thermal-test": {
        "columns": ["id","site_id","test_pit_id","depth","thermal_reading","r_value"],
        "required": ["id","test_pit_id","depth","thermal_reading","r_value"],
        "desc": "In-situ thermal resistivity tests.",
    },
    "lab-test": {
        "columns": ["id","site_id","location_id","soil_unit_no","top_depth","bottom_depth",
                    "moisture_content","liquid_limit","plastic_limit","plasticity_index",
                    "linear_shrinkage","emerson_class","iss","gravel","sand","fines",
                    "compaction_mdd","compaction_omc","cbr_4day_2_5mm","cbr_swell"],
        "required": ["id","location_id"],
        "desc": "Laboratory test results linked to a borehole or test pit.",
    },
    "aggressivity": {
        "columns": ["id","site_id","location_id","depth","ph","sulfate","chlorides",
                    "resistivity","exposure_class_concrete","exposure_class_steel"],
        "required": ["id","location_id"],
        "desc": "Soil aggressivity / corrosivity tests.",
    },
    "pile-test-location": {
        "columns": ["id","site_id","zone_id","driving_type","easting","northing","reduced_level",
                    "designer","section_type","target_depth","achieved_embedment",
                    "drive_time","driving_rate"],
        "required": ["id"],
        "desc": "Pile locations with install data. driving_type = Driven | PreDrilled.",
    },
    "pile-test": {
        "columns": ["id","site_id","pile_location_id","section_type","passed"],
        "required": ["id","pile_location_id"],
        "desc": "Pile test container. passed = true | false | undecided.",
    },
    "tension-test": {
        "columns": ["id","site_id","pile_test_id","uplift_applied_force",
                    "uplift_max_deflection","max_load_proportion_ed"],
        "required": ["id","pile_test_id"],
        "desc": "Tension / uplift sub-test.",
    },
    "lateral-test": {
        "columns": ["id","site_id","pile_test_id","max_applied_force","max_deflection_top",
                    "max_deflection_bottom","load_max","max_load_proportion_ed"],
        "required": ["id","pile_test_id"],
        "desc": "Lateral sub-test.",
    },
    "compression-test": {
        "columns": ["id","site_id","pile_test_id","max_applied_force",
                    "max_deflection","max_load_proportion_ed"],
        "required": ["id","pile_test_id"],
        "desc": "Compression sub-test.",
    },
}

SCHEMA_SUMMARY = "\n".join(
    f"  {name}: {', '.join(s['columns'])}  [{s['desc']}]"
    for name, s in SCHEMA.items()
)


# ── Column alias map — maps contractor terminology to schema field names ───────
# Deliberately conservative: only map when the alias is unambiguous.
# Unknown columns are flagged for the LLM rather than guessed.

COLUMN_ALIASES: dict[str, dict[str, str]] = {
    "soil-type": {
        "unit no":         "unit_no",
        "unit number":     "unit_no",
        "no.":             "unit_no",
        "no":              "unit_no",
        "origin":          "origin",
        "formation":       "origin",
        "geological unit": "origin",
        "unit name":       "unit_name",
        "name":            "unit_name",
        "material":        "unit_name",
        "soil name":       "unit_name",
        "description":     "description",
        "material description": "description",
    },
    "borehole": {
        "borehole id":    "id",
        "borehole":       "id",
        "bh":             "id",
        "hole id":        "id",
        "easting":        "easting",
        "e (m)":          "easting",
        "northing":       "northing",
        "n (m)":          "northing",
        "rl":             "elevation",
        "elevation":      "elevation",
        "reduced level":  "elevation",
        "depth":          "total_depth",
        "total depth":    "total_depth",
        "depth (m)":      "total_depth",
        "groundwater":    "groundwater_depth",
        "gwl":            "groundwater_depth",
        "water table":    "groundwater_depth",
        "series":         "series",
    },
    "testpit": {
        "test pit":       "id",
        "trial pit":      "id",
        "tp":             "id",
        "pit id":         "id",
        "easting":        "easting",
        "northing":       "northing",
        "rl":             "elevation",
        "elevation":      "elevation",
        "depth":          "total_depth",
        "total depth":    "total_depth",
    },
    "dpsh": {
        "probe id":       "id",
        "dpsh":           "id",
        "id":             "id",
        "easting":        "easting",
        "northing":       "northing",
        "refusal depth":  "refusal_depth",
        "refusal":        "refusal_depth",
        "depth of refusal": "refusal_depth",
        "depth (m)":      "refusal_depth",
    },
    "thermal-test": {
        "location":       "test_pit_id",
        "test pit":       "test_pit_id",
        "tp":             "test_pit_id",
        "depth":          "depth",
        "depth (m)":      "depth",
        "thermal reading": "thermal_reading",
        "k (w/mk)":       "thermal_reading",
        "k(w/mk)":        "thermal_reading",
        "thermal resistivity": "thermal_reading",
        "r-value":        "r_value",
        "r value":        "r_value",
        "r (ccm/w)":      "r_value",
    },
    "lab-test": {
        "location":       "location_id",
        "borehole":       "location_id",
        "sample":         "location_id",
        "top depth":      "top_depth",
        "from (m)":       "top_depth",
        "from":           "top_depth",
        "bottom depth":   "bottom_depth",
        "to (m)":         "bottom_depth",
        "to":             "bottom_depth",
        "material encountered": "material",
        "material":       "material",
        "moisture content": "moisture_content",
        "mc (%)":         "moisture_content",
        "liquid limit":   "liquid_limit",
        "ll (%)":         "liquid_limit",
        "ll":             "liquid_limit",
        "plastic limit":  "plastic_limit",
        "pl (%)":         "plastic_limit",
        "pl":             "plastic_limit",
        "plasticity index": "plasticity_index",
        "pi (%)":         "plasticity_index",
        "pi":             "plasticity_index",
        "linear shrinkage": "linear_shrinkage",
        "ls (%)":         "linear_shrinkage",
        "emerson class":  "emerson_class",
        "emerson":        "emerson_class",
        "iss":            "iss",
        "gravel (%)":     "gravel",
        "gravel":         "gravel",
        "sand (%)":       "sand",
        "sand":           "sand",
        "fines (%)":      "fines",
        "fines":          "fines",
        "mdd":            "compaction_mdd",
        "mdd (t/m3)":     "compaction_mdd",
        "omc":            "compaction_omc",
        "omc (%)":        "compaction_omc",
        "cbr":            "cbr_4day_2_5mm",
        "@2.5mm":         "cbr_4day_2_5mm",
        "swell":          "cbr_swell",
        "specimen swell": "cbr_swell",
    },
    "aggressivity": {
        "location":       "location_id",
        "borehole":       "location_id",
        "depth":          "depth",
        "depth (m bgl)":  "depth",
        "ph":             "ph",
        "sulfate":        "sulfate",
        "sulfate content": "sulfate",
        "sulphate":       "sulfate",
        "chlorides":      "chlorides",
        "chloride":       "chlorides",
        "resistivity":    "resistivity",
        "exposure classification concrete": "exposure_class_concrete",
        "concrete piles": "exposure_class_concrete",
        "exposure classification steel":    "exposure_class_steel",
        "steel piles":    "exposure_class_steel",
    },
    "pile-test-location": {
        "pile id":        "id",
        "pile":           "id",
        "location":       "id",
        "loc":            "id",
        "plt":            "id",
        "zone":           "zone_id",
        "block":          "zone_id",
        "pcu":            "zone_id",
        "easting":        "easting",
        "northing":       "northing",
        "rl":             "reduced_level",
        "reduced level":  "reduced_level",
        "section type":   "section_type",
        "section":        "section_type",
        "target depth":   "target_depth",
        "design depth":   "target_depth",
        "achieved embedment": "achieved_embedment",
        "achieved":       "achieved_embedment",
        "embedment":      "achieved_embedment",
        "drive time":     "drive_time",
        "driving time":   "drive_time",
        "driving rate":   "driving_rate",
        "rate":           "driving_rate",
        "driving type":   "driving_type",
        "type":           "driving_type",
        "designer":       "designer",
    },
}


def _map_headers(table: "ExtractedTable", sheet_name: str) -> dict[str, str]:
    """
    Map table column headers to schema field names using the alias dictionary.
    Returns {table_header: schema_field} for matched columns only.
    Unmatched columns are excluded and flagged in the LLM prompt.
    """
    aliases = COLUMN_ALIASES.get(sheet_name, {})
    mapping: dict[str, str] = {}
    for header in table.headers:
        key = header.lower().strip()
        if key in aliases:
            mapping[header] = aliases[key]
    return mapping


# ── Targeted LLM prompts per table type ───────────────────────────────────────

# Maps heuristic table_type → schema sheet names it could produce
TABLE_TYPE_TO_SHEETS: dict[str, list[str]] = {
    "soil_profile":     ["soil-type"],
    "ground_model":     ["ground-layer", "ground-model"],
    "borehole_summary": ["borehole"],
    "testpit_summary":  ["testpit"],
    "laboratory_test":  ["lab-test"],
    "thermal_test":     ["thermal-test"],
    "aggressivity":     ["aggressivity"],
    "pile_test":        ["pile-test-location", "pile-test",
                         "tension-test", "lateral-test", "compression-test"],
    "dpsh":             ["dpsh"],
}


def _targeted_prompt(table: "ExtractedTable") -> str:
    """
    Build a targeted LLM prompt for a single classified table.
    Includes the column mapping found by the algorithm, flags unmapped columns,
    and tells the LLM what sheet(s) to produce.
    """
    t_type    = table.table_type
    sheets    = TABLE_TYPE_TO_SHEETS.get(t_type, [])
    schema_lines = "\n".join(
        f"  {name}: {', '.join(SCHEMA[name]['columns'])}  [{SCHEMA[name]['desc']}]"
        for name in sheets if name in SCHEMA
    )

    # Column mapping from algorithm
    primary_sheet = sheets[0] if sheets else None
    mapping       = _map_headers(table, primary_sheet) if primary_sheet else {}
    mapped_str    = "\n".join(
        f"  '{col}' → {field}" for col, field in mapping.items()
    ) or "  (no automatic column matches found)"
    unmapped = [h for h in table.headers if h not in mapping]
    unmapped_str = ", ".join(f"'{c}'" for c in unmapped) if unmapped else "none"

    context = ""
    if table.title:
        context = f"Table title from document: {table.title}\n"
    if table.raw_text_above:
        context += f"Context above table: {table.raw_text_above[:200]}\n"

    return f"""You are a geotechnical data extractor. Extract data from the table below.
Return ONLY a valid JSON object mapping sheet names to arrays of row objects.

{context}
TARGET SCHEMA (only produce these sheets):
{schema_lines}

COLUMN MAPPING (already resolved by algorithm — use these):
{mapped_str}

UNMAPPED COLUMNS (use your judgement — map only if confident, otherwise null):
{unmapped_str}

RULES:
- Use the column mapping above. Do not second-guess mapped columns.
- For unmapped columns: match by meaning only if very confident. If uncertain → null.
- Use null for missing/blank values. Never invent data.
- IDs: use the exact label from the document (e.g. BH01, TP03, PLT-004A).
- Numbers: strip units, store as numbers (1.65 not "1.65 m").
- zone_id: always null (assigned separately from site map).
- Do not include rows that are entirely null.

TABLE ({table.row_count()} rows, {table.column_count()} columns):
{table.as_text()}

Return JSON only. No markdown, no explanation."""


def _generic_prompt(table: "ExtractedTable") -> str:
    """
    Fallback prompt for unclassified tables — uses the full schema
    but explicitly warns against over-eager mapping.
    """
    context = f"Table title: {table.title}\n" if table.title else ""
    context += f"Context: {table.raw_text_above[:200]}\n" if table.raw_text_above else ""

    return f"""You are a geotechnical data extractor. A table from a PDF report is shown below.
Determine which schema sheet(s) it belongs to and extract the data.
Return ONLY a valid JSON object, or an empty object {{}} if no data can be extracted.

{context}
FULL SCHEMA:
{SCHEMA_SUMMARY}

RULES:
- Map to a schema sheet ONLY if you are confident — do not force a fit.
- If the table doesn't match any schema sheet, return {{}}.
- Column names may differ from schema names — map by meaning, not by name.
- Be conservative: uncertain columns → null rather than a guess.
- null for missing values. Numbers without units. zone_id always null.

TABLE:
{table.as_text()}

Return JSON only."""


# ── Merge results ─────────────────────────────────────────────────────────────

# ── Merge results from multiple PDFs ─────────────────────────────────────────

def _merge(results: list[dict]) -> dict:
    merged: dict[str, list] = {k: [] for k in SCHEMA}
    seen: dict[str, set] = {k: set() for k in SCHEMA}
    for r in results:
        for sheet, rows in r.items():
            if sheet not in merged or not isinstance(rows, list):
                continue
            id_col = "unit_no" if sheet == "soil-type" else "id"
            for row in rows:
                if not isinstance(row, dict):
                    continue
                key = str(row.get(id_col, ""))
                if key and key in seen[sheet]:
                    continue  # deduplicate by id
                if key:
                    seen[sheet].add(key)
                merged[sheet].append(row)
    return {k: v for k, v in merged.items() if v}


# ── Build xlsx from merged data ───────────────────────────────────────────────

def _build_xlsx(data: dict) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment

    HFONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    HFILL  = PatternFill("solid", start_color="2F6F77")
    RFILL  = PatternFill("solid", start_color="FFE0E0")   # required
    RFONT  = Font(bold=True, color="CC0000", name="Arial", size=10)
    FKFILL = PatternFill("solid", start_color="FFF3CC")   # foreign key
    FKFONT = Font(bold=True, color="7A5200", name="Arial", size=10)
    NULL_FILL = PatternFill("solid", start_color="F5F5F5")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, schema in SCHEMA.items():
        ws = wb.create_sheet(sheet_name)
        cols = schema["columns"]
        req  = set(schema.get("required", []))
        fk_cols = {c for c in cols if c.endswith("_id")}

        ws.append(cols)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(1, ci)
            if col in req:
                cell.font = RFONT; cell.fill = RFILL
            elif col in fk_cols:
                cell.font = FKFONT; cell.fill = FKFILL
            else:
                cell.font = HFONT; cell.fill = HFILL
            cell.alignment = Alignment(horizontal="center")

        rows = data.get(sheet_name, [])
        for row in rows:
            ws.append([row.get(c) for c in cols])

        # grey out null cells so they're visually distinct from blanks
        for r in ws.iter_rows(min_row=2):
            for c in r:
                if c.value is None:
                    c.fill = NULL_FILL

        # auto-width
        for col in ws.columns:
            w = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = max(w, 12)

        ws.freeze_panes = "A2"

        # legend in top-right
        legend_col = len(cols) + 2
        legend = [
            ("■ RED header", "CC0000", "Required field"),
            ("■ ORANGE header", "7A5200", "Foreign key (links to another sheet)"),
            ("■ TEAL header", "2F6F77", "Optional field"),
            ("■ Grey cell", "888888", "No data extracted — fill manually"),
        ]
        for i, (label, color, note) in enumerate(legend, 1):
            ws.cell(i, legend_col).value = f"{label}: {note}"
            ws.cell(i, legend_col).font = Font(color=color, italic=True, size=9, name="Arial")

    return io.BytesIO(b"").getvalue().__class__(wb.save(io.BytesIO()) or b"") or _save(wb)


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Routes ────────────────────────────────────────────────────────────────────

@router.delete("/extract/{job_id}")
def cancel_extract(job_id: str):
    """Signal a running extraction job to stop."""
    ev = _cancel_events.get(job_id)
    if ev:
        ev.set()
        return {"cancelled": True}
    return {"cancelled": False}


def _sse(event: str, data: dict) -> str:
    return f"event: {event}\ndata: {json.dumps(data)}\n\n"


def _run_extraction(files_data: list, cancel: threading.Event, job_id: str):
    """
    Two-stage extraction generator yielding SSE strings.

    Stage 1 — Algorithm:
        pdfplumber extracts tables, captures titles/context,
        heuristics classify each table by type, column aliases
        map known headers to schema fields.

    Stage 2 — LLM:
        Each classified table gets a targeted prompt containing
        only the relevant schema sheets and the algorithm's column
        mapping. Unclassified tables get a conservative generic prompt.
        The LLM maps remaining/ambiguous columns and fills values.
    """
    file_plans = []
    for filename, data in files_data:
        extracted = extract_pdf(data)
        tables = extracted["extracted_tables"]
        if not tables:
            yield _sse("warning", {"message": f"{filename}: no tables found, skipped"})
            continue
        file_plans.append((filename, tables))

    if not file_plans:
        yield _sse("error", {"message": "No tables found in any uploaded PDF"})
        return

    total_chunks = sum(len(tables) for _, tables in file_plans)
    yield _sse("start", {"total_chunks": total_chunks,
                          "total_files": len(file_plans)})

    done_chunks = 0
    all_results, all_errors = [], []

    for filename, tables in file_plans:
        if cancel.is_set():
            yield _sse("cancelled", {"message": "Extraction cancelled"})
            return

        # Log classification summary
        from collections import Counter
        types = Counter(t.table_type or "unclassified" for t in tables)
        logger.info("%s: %d tables — %s", filename,
                    len(tables), dict(types))

        file_result: dict = {}

        for table in tables:
            if cancel.is_set():
                yield _sse("cancelled", {"message": "Extraction cancelled"})
                return

            label = (f"{filename} p{table.page} "
                     f"table {table.table_index}"
                     + (f" [{table.table_type}]" if table.table_type else " [unclassified]"))

            yield _sse("chunk_start", {
                "file":          filename,
                "chunk":         done_chunks + 1,
                "done":          done_chunks,
                "total":         total_chunks,
                "label":         label,
                "table_type":    table.table_type,
                "table_title":   table.title,
                "n_rows":        table.row_count(),
                "n_cols":        table.column_count(),
            })

            # Skip tiny tables — likely artefacts or formatting
            if table.row_count() < 1 or table.column_count() < 2:
                done_chunks += 1
                yield _sse("chunk_done", {"done": done_chunks, "total": total_chunks,
                                           "skipped": True, "reason": "too small"})
                continue

            # Build targeted or generic prompt
            if table.table_type:
                prompt = _targeted_prompt(table)
            else:
                prompt = _generic_prompt(table)

            try:
                raw    = _llm_extract(prompt, cancel=cancel)
                parsed = _parse_json(raw)
                if isinstance(parsed, dict) and parsed:
                    for sheet, rows in parsed.items():
                        if isinstance(rows, list) and rows:
                            file_result.setdefault(sheet, []).extend(rows)
            except RuntimeError as e:
                msg = str(e)
                if "Cancelled" in msg:
                    yield _sse("cancelled", {"message": "Extraction cancelled"})
                    return
                if "rate" in msg.lower():
                    yield _sse("rate_limit", {
                        "message": "Rate limited — waiting before retry…",
                        "chunk": label,
                    })
                all_errors.append(f"{label}: {msg}")
            except Exception as ex:
                all_errors.append(f"{label}: {ex}")

            done_chunks += 1
            yield _sse("chunk_done", {"done": done_chunks, "total": total_chunks})

            # Pause between LLM calls
            if done_chunks < total_chunks:
                pause = 3 if total_chunks > 10 else 1
                for _ in range(pause):
                    if cancel.is_set():
                        yield _sse("cancelled", {"message": "Extraction cancelled"})
                        return
                    time.sleep(1)

        if file_result:
            all_results.append(file_result)

    if not all_results:
        yield _sse("error", {"message": "No data extracted", "errors": all_errors})
        return

    merged  = _merge(all_results)
    summary = {k: len(v) for k, v in merged.items()}
    yield _sse("done", {"data": merged, "summary": summary, "errors": all_errors})



@router.post("/extract/{job_id}")
async def extract_pdfs(job_id: str, request: Request,
                       files: List[UploadFile] = File(...)):
    """
    SSE streaming extraction. Cancel via DELETE /ingest/extract/{job_id}.
    Events: start | chunk_start | chunk_done | rate_limit | warning | done | error | cancelled
    """
    if not files:
        raise HTTPException(400, "No files uploaded")

    files_data = []
    for f in files:
        if f.filename.lower().endswith(".pdf"):
            files_data.append((f.filename, await f.read()))

    cancel = threading.Event()
    _cancel_events[job_id] = cancel

    def generate():
        try:
            yield from _run_extraction(files_data, cancel, job_id)
        finally:
            _cancel_events.pop(job_id, None)

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.get("/template")
def download_template():
    """Return the blank Example_Input template xlsx."""
    blank = _build_xlsx({})
    return StreamingResponse(
        io.BytesIO(blank),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="geotech_input_template.xlsx"'},
    )


@router.post("/download-prefilled")
async def download_prefilled(files: List[UploadFile] = File(...)):
    """Extract from PDFs and return a pre-filled xlsx for the user to edit."""
    resp = await extract_pdfs(files)
    xlsx = _build_xlsx(resp["data"])
    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="geotech_extracted.xlsx"'},
    )


@router.post("/import")
async def import_xlsx(file: UploadFile = File(...)):
    """
    Accept the final edited xlsx and bulk-write every sheet to the graph
    via the existing /nodes/{type}/bulk endpoints (internal call).
    """
    import httpx, os
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Expected an .xlsx file")

    data    = await file.read()
    wb      = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    base    = os.getenv("INTERNAL_API_BASE", "http://localhost:8000")

    results = {}
    async with httpx.AsyncClient(base_url=base, timeout=60) as client:
        for sheet_name in wb.sheetnames:
            if sheet_name not in SCHEMA:
                continue
            ws   = wb[sheet_name]
            cols = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                obj = {cols[i]: v for i, v in enumerate(row) if i < len(cols)}
                # skip rows where the id field is blank
                id_col = "unit_no" if sheet_name == "soil-type" else "id"
                if not obj.get(id_col):
                    continue
                # convert Excel booleans / strings
                for k, v in obj.items():
                    if isinstance(v, str) and v.lower() in ("true","false"):
                        obj[k] = v.lower() == "true"
                rows.append(obj)

            if not rows:
                continue

            try:
                r = await client.post(f"/nodes/{sheet_name}/bulk", json={"rows": rows})
                results[sheet_name] = r.json()
            except Exception as e:
                results[sheet_name] = {"error": str(e)}

    return {"imported": results}