"""Parse the pre-drill block-calculation spreadsheet into Zone (block) records.

Input columns (Sheet1):
    Block No. | Predrill/Driven | 4-String Tracker | 3-String Tracker | 2-String Tracker

Each data row becomes a Zone with id ZONE-<block> carrying the pre-drill decision
and tracker counts. Returns plain dicts (matching the /nodes/zone/upsert schema);
nothing is pushed to the graph here — that's the caller's job (see ingest_maryvale).

Usage:
    from app.ingestion.parse_predrill import parse_predrill
    zones = parse_predrill("/path/Pre-drill_block_calculation.xlsx")
"""
from openpyxl import load_workbook


def _block_id(raw) -> str:
    # Excel stores "1.1" as the float 1.1000000000000001; render cleanly.
    if isinstance(raw, (int, float)):
        return f"ZONE-{raw:.1f}"
    return f"ZONE-{str(raw).strip()}"


def parse_predrill(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    zones, header_seen = [], False
    for row in ws.iter_rows(values_only=True):
        if row is None or all(c is None for c in row):
            continue
        first = row[0]
        # locate the header row, then read data rows until the totals
        if not header_seen:
            if first and str(first).strip().lower().startswith("block"):
                header_seen = True
            continue
        # stop at summary rows ("PRE DRILL TOTAL", "DRIVEN TOTAL")
        if isinstance(first, str) and "total" in first.lower():
            continue
        if not isinstance(first, (int, float)):
            continue

        decision = (str(row[1]).strip() if row[1] is not None else None)
        zones.append({
            "id": _block_id(first),
            "name": f"Block {first:.1f}" if isinstance(first, (int, float)) else f"Block {first}",
            "pre_drill_decision": decision,
            "trackers_4string": int(row[2]) if row[2] is not None else None,
            "trackers_3string": int(row[3]) if row[3] is not None else None,
            "trackers_2string": int(row[4]) if row[4] is not None else None,
        })
    wb.close()
    return zones


if __name__ == "__main__":
    import json, sys
    rows = parse_predrill(sys.argv[1])
    print(f"{len(rows)} blocks parsed")
    print(json.dumps(rows[:3], indent=2))